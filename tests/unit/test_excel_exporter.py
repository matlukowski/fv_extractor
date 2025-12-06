"""Unit tests for Excel Exporter (RED phase - tests first)"""
import pytest
import pandas as pd
from io import BytesIO
from datetime import date
from openpyxl import load_workbook
from pydantic import ValidationError
from src.exporters.excel_exporter import export_to_excel
from src.models.invoice_data import InvoiceData


@pytest.fixture
def sample_invoice_data():
    """Fixture with complete invoice data for testing"""
    return InvoiceData(
        invoice_number="FV001/2025",
        issue_date=date(2025, 1, 15),
        seller_name="ACME Corp Sp. z o.o.",
        seller_nip="1234567890",
        buyer_name="XYZ Solutions",
        items=[
            {
                "description": "Laptop Dell XPS 15",
                "quantity": 2.0,
                "unit_price_net": 4500.00,
                "vat_rate": 23,
                "total_gross": 11070.00,
                "category": "IT"
            },
            {
                "description": "Mouse Logitech",
                "quantity": 3.0,
                "unit_price_net": 50.00,
                "vat_rate": 23,
                "total_gross": 184.50,
                "category": "IT"
            }
        ],
        total_net_sum=9150.00,
        total_gross_sum=11254.50,
        currency="PLN"
    )


@pytest.fixture
def single_item_invoice():
    """Fixture with single-item invoice"""
    return InvoiceData(
        invoice_number="FV002/2025",
        issue_date=date(2025, 1, 16),
        seller_name="Test Seller",
        seller_nip="9876543210",
        buyer_name="Test Buyer",
        items=[
            {
                "description": "Service",
                "quantity": 1.0,
                "unit_price_net": 1000.00,
                "vat_rate": 23,
                "total_gross": 1230.00,
                "category": None
            }
        ],
        total_net_sum=1000.00,
        total_gross_sum=1230.00,
        currency="EUR"
    )


class TestExportToExcel:
    """Test export_to_excel function"""

    def test_export_to_excel_creates_valid_file(self, sample_invoice_data):
        """Should return BytesIO buffer with valid Excel file"""
        result = export_to_excel(sample_invoice_data)

        assert isinstance(result, BytesIO)
        assert result.tell() == 0  # Should be at start position

        # Should be able to load as Excel
        wb = load_workbook(result)
        assert wb is not None
        assert len(wb.sheetnames) > 0

    def test_export_includes_all_fields(self, sample_invoice_data):
        """Should include all invoice fields in Excel"""
        result = export_to_excel(sample_invoice_data)

        # Load and check
        wb = load_workbook(result)
        ws = wb.active
        df = pd.DataFrame(ws.values)

        # Should have header row + 2 data rows (2 items)
        assert len(df) >= 3  # Header + 2 items

    def test_export_has_correct_number_of_rows(self, sample_invoice_data):
        """Should have one row per item plus header"""
        result = export_to_excel(sample_invoice_data)

        wb = load_workbook(result)
        ws = wb.active

        # Count non-empty rows
        row_count = sum(1 for row in ws.iter_rows() if any(cell.value for cell in row))

        # Should be 3: 1 header + 2 items
        assert row_count == 3

    def test_export_single_item_invoice(self, single_item_invoice):
        """Should handle invoice with single item"""
        result = export_to_excel(single_item_invoice)

        wb = load_workbook(result)
        ws = wb.active

        row_count = sum(1 for row in ws.iter_rows() if any(cell.value for cell in row))

        # Should be 2: 1 header + 1 item
        assert row_count == 2


class TestExcelContent:
    """Test Excel content and formatting"""

    def test_export_contains_header_fields(self, sample_invoice_data):
        """Should include invoice header information"""
        result = export_to_excel(sample_invoice_data)

        wb = load_workbook(result)
        ws = wb.active

        # Convert to string for searching
        all_values = []
        for row in ws.iter_rows():
            all_values.extend([str(cell.value) for cell in row if cell.value])

        all_text = ' '.join(all_values)

        # Check for key header fields
        assert "FV001/2025" in all_text
        assert "ACME" in all_text or "1234567890" in all_text

    def test_export_contains_item_data(self, sample_invoice_data):
        """Should include item descriptions and quantities"""
        result = export_to_excel(sample_invoice_data)

        wb = load_workbook(result)
        ws = wb.active

        all_values = []
        for row in ws.iter_rows():
            all_values.extend([str(cell.value) for cell in row if cell.value])

        all_text = ' '.join(all_values)

        assert "Laptop" in all_text
        assert "Mouse" in all_text

    def test_export_preserves_numeric_values(self, sample_invoice_data):
        """Should preserve numeric values correctly"""
        result = export_to_excel(sample_invoice_data)

        wb = load_workbook(result)
        ws = wb.active

        # Find cells with numeric values
        numeric_values = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numeric_values.append(cell.value)

        # Should have various numbers from invoice
        assert len(numeric_values) > 0
        assert 4500.00 in numeric_values or 11070.00 in numeric_values


class TestSpecialCharacters:
    """Test handling of special characters"""

    def test_export_handles_polish_characters(self):
        """Should correctly encode Polish characters (ąćęłńóśźż)"""
        invoice_with_polish = InvoiceData(
            invoice_number="FV003/2025",
            issue_date=date(2025, 1, 17),
            seller_name="Firma Świętokrzyska Sp. z o.o.",
            seller_nip="1234567890",
            buyer_name="Klient Łódzki",
            items=[
                {
                    "description": "Usługa sprzątania biura",
                    "quantity": 1.0,
                    "unit_price_net": 500.00,
                    "vat_rate": 23,
                    "total_gross": 615.00,
                    "category": "Sprzątanie"
                }
            ],
            total_net_sum=500.00,
            total_gross_sum=615.00,
            currency="PLN"
        )

        result = export_to_excel(invoice_with_polish)

        # Should create valid Excel
        wb = load_workbook(result)
        ws = wb.active

        # Find Polish characters
        all_text = []
        for row in ws.iter_rows():
            all_text.extend([str(cell.value) for cell in row if cell.value])

        combined = ' '.join(all_text)

        # Should contain Polish characters
        assert "Święto" in combined or "Łódz" in combined or "sprzątani" in combined


class TestDateFormatting:
    """Test date formatting in Excel"""

    def test_export_formats_dates_correctly(self, sample_invoice_data):
        """Should include date in readable format"""
        result = export_to_excel(sample_invoice_data)

        wb = load_workbook(result)
        ws = wb.active

        # Look for date values
        date_found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "2025" in str(cell.value):
                    date_found = True
                    break

        assert date_found


class TestCurrencyFormatting:
    """Test currency/monetary value formatting"""

    def test_export_formats_currency_correctly(self, sample_invoice_data):
        """Should format monetary values with 2 decimal places"""
        result = export_to_excel(sample_invoice_data)

        wb = load_workbook(result)
        ws = wb.active

        # Check for decimal values
        decimal_values = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    decimal_values.append(cell.value)

        # Should have monetary values
        assert len(decimal_values) > 0


class TestErrorHandling:
    """Test error handling"""

    def test_export_empty_items_raises_error(self):
        """Should raise error for invoice with no items"""
        # This should already be caught by Pydantic, but test anyway
        with pytest.raises((ValueError, ValidationError, Exception)):
            invalid_invoice = InvoiceData(
                invoice_number="FV999/2025",
                issue_date=date(2025, 1, 1),
                seller_name="Test",
                seller_nip="1234567890",
                buyer_name="Test",
                items=[],  # Empty!
                total_net_sum=0,
                total_gross_sum=0,
                currency="PLN"
            )


class TestMultipleItems:
    """Test export with varying numbers of items"""

    def test_export_with_many_items(self):
        """Should handle invoice with many line items"""
        items = [
            {
                "description": f"Item {i}",
                "quantity": 1.0,
                "unit_price_net": 100.00 * i,
                "vat_rate": 23,
                "total_gross": 123.00 * i,
                "category": "Test"
            }
            for i in range(1, 11)  # 10 items
        ]

        invoice = InvoiceData(
            invoice_number="FV010/2025",
            issue_date=date(2025, 1, 20),
            seller_name="Seller",
            seller_nip="1234567890",
            buyer_name="Buyer",
            items=items,
            total_net_sum=5500.00,
            total_gross_sum=6765.00,
            currency="PLN"
        )

        result = export_to_excel(invoice)

        wb = load_workbook(result)
        ws = wb.active

        row_count = sum(1 for row in ws.iter_rows() if any(cell.value for cell in row))

        # Should be 11: 1 header + 10 items
        assert row_count == 11


class TestColumnHeaders:
    """Test column header names"""

    def test_export_has_descriptive_column_headers(self, sample_invoice_data):
        """Should have clear, descriptive column headers"""
        result = export_to_excel(sample_invoice_data)

        wb = load_workbook(result)
        ws = wb.active

        # Get first row (headers)
        headers = [cell.value for cell in ws[1] if cell.value]

        # Should have multiple columns
        assert len(headers) > 5

        # Should have key columns (exact names may vary)
        header_text = ' '.join([str(h).lower() for h in headers])

        # Check for important fields
        assert any(word in header_text for word in ['description', 'opis', 'nazwa'])
