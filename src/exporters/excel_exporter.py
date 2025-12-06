"""Excel exporter for invoice data"""
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..models.invoice_data import InvoiceData


def export_to_excel(invoice_data: InvoiceData) -> BytesIO:
    """
    Export InvoiceData to Excel file (.xlsx) in memory.

    Creates a formatted Excel spreadsheet with:
    - Header row with column names
    - One row per invoice line item
    - All invoice metadata repeated for each row
    - Formatted cells (bold headers, aligned columns)

    Args:
        invoice_data: Validated InvoiceData object

    Returns:
        BytesIO buffer containing Excel file (ready for download/save)

    Raises:
        ValueError: If invoice has no items

    Examples:
        >>> invoice = InvoiceData(...)
        >>> excel_buffer = export_to_excel(invoice)
        >>> with open('invoice.xlsx', 'wb') as f:
        ...     f.write(excel_buffer.read())
    """
    if not invoice_data.items:
        raise ValueError("Invoice must have at least one item")

    # Convert invoice data to DataFrame using InvoiceData's method
    df = invoice_data.to_dataframe()

    # Reorder columns for better readability
    column_order = [
        'invoice_number',
        'issue_date',
        'seller_name',
        'seller_nip',
        'buyer_name',
        'description',
        'quantity',
        'unit_price_net',
        'vat_rate',
        'total_gross',
        'category',
        'currency',
        'total_net_sum',
        'total_gross_sum'
    ]

    # Only include columns that exist
    available_columns = [col for col in column_order if col in df.columns]
    df = df[available_columns]

    # Rename columns to human-readable Polish names
    column_names = {
        'invoice_number': 'Numer faktury',
        'issue_date': 'Data wystawienia',
        'seller_name': 'Sprzedawca',
        'seller_nip': 'NIP sprzedawcy',
        'buyer_name': 'Nabywca',
        'description': 'Opis pozycji',
        'quantity': 'Ilość',
        'unit_price_net': 'Cena jedn. netto',
        'vat_rate': 'VAT %',
        'total_gross': 'Wartość brutto',
        'category': 'Kategoria',
        'currency': 'Waluta',
        'total_net_sum': 'Suma netto',
        'total_gross_sum': 'Suma brutto'
    }

    df = df.rename(columns={k: v for k, v in column_names.items() if k in df.columns})

    # Create Excel file in memory
    output = BytesIO()

    # Use openpyxl engine for better formatting control
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Faktura')

        # Get workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Faktura']

        # Format header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:  # First row (header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set column width (with padding)
            adjusted_width = min(max_length + 2, 50)  # Max 50 chars
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Format numeric columns
        for row in worksheet.iter_rows(min_row=2):  # Skip header
            for cell in row:
                # Align numbers to the right
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')

                    # Format currency columns with 2 decimal places
                    if cell.column_letter in ['H', 'J', 'M', 'N']:  # Price columns
                        cell.number_format = '#,##0.00'

    # Reset buffer position to beginning
    output.seek(0)

    return output
